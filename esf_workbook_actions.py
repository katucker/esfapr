# -*- coding: utf-8 -*-
"""ESF APR Excel workbook functions.

Python module for buidling data objects from Excel workbooks for
Annual Performance Reports for the Education Stabilization Fund
grants.

@author: Keith.Tucker
"""
from collections.abc import Iterable
from dataclasses import dataclass
import logging
from typing import List, Any

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

class _ESF_Sub:
    """Child class for all subordinate pieces of an APR.

    This class is used as the common building block for all
    subordinate pieces of an APR, such as a list of subawards
    or a list of schools receiving benefits.
    Since there are no common data elements in the subordinate
    pieces, this class is defined with no attributes.
    """

    def merge(self, other_instance):
        """Add all the attributes in the other_instance object to this instance.
        Contrary to the update method implemented on some Python types, this
        method does not overwrite existing attributes, it only adds new ones."""
        for attr_name in vars(other_instance):
            if hasattr(self, attr_name):
                continue
            setattr(self, attr_name, getattr(other_instance,attr_name))


@dataclass
class ESF_APR:
    """Parent class for generalizing functions that use APR data parameters.
    
    This class contains data elements common to all APRs, including
    ones that may not be present in the data file, such as the OMB 
    control number displayed on the associated APR form.
    Other attributes specific to each type and version of APR are
    added dynamically after creation of an instance of this class.
    """
    omb_control_number: str
    expiration_date: str
    cui_official: str
    reporting_year: str


def _empty_gen():
    """Empty iterator generator function."""
    yield from ()


def _key_gen(keys: List[str]):
    """Iterate over a list of key strings."""
    yield from keys


def _force_bool(input_value: Any) -> bool:
    """Coerce a value to boolean type."""
    if isinstance(input_value,bool):
        return input_value
    string_value = str(input_value)
    return string_value.lower() == "true"


def _force_float(input_value: Any) -> float:
    """Coerce a value to float type."""
    if isinstance(input_value,float):
        return input_value
    try:
        float_value = float(input_value)
        return float_value
    except (TypeError, ValueError):
        return 0.0


def _force_int(input_value: Any) -> int:
    """Coerce a value to int type."""
    if isinstance(input_value,int):
        return input_value
    try:
        int_value = int(input_value)
        return int_value
    except (TypeError, ValueError):
        return 0


def _find_row(ws: Worksheet, key: str, key_column: int) -> int:
    """Find a row within a worksheet in which the passed key equals the value in the passed key_column."""
    for row in ws.iter_rows(min_row=2, min_col=key_column, max_col=key_column):
        # Each row is a list containing just the cell for the key_column.
        cell = row[0]
        if cell.value == key:
            return cell.row
    # Fallback if key is not found
    return -1


def _extract_sub_worksheet(ws: Worksheet, workbook_map: list, key: str, key_offset: int) -> List[_ESF_Sub]:
    subs = []
    try:
        logging.info(f'Iterating {ws.title} for key {key}')
        for row in ws.iter_rows(min_row=2, min_col=1,
                                max_col=len(workbook_map), values_only=True):
            if row[key_offset] == key:
                sub = _ESF_Sub()
                # Loop over the entries in the workbook_map, creating an attribute on
                # the sub object matching the name in the workbook_map entry, using the
                # value at the index in the workbook_map entry, and matching the type
                # specified in the workbook_map entry.
                for element in workbook_map:
                    index = element.get('index', None)
                    if index is None:
                        logging.info(f'Missing index value in {element}')
                        continue
                    if index > len(row):
                        logging.info(
                            f'Index {index} exceeds row tuple length of {len(row)} for key {key}')
                        continue
                    attr_name = element.get('name', None)
                    if attr_name is None:
                        logging.info(f'Missing attribute name in {element}')
                        continue
                    match element.get('type', None):
                        case 'bool':
                            val = _force_bool(row[index])
                        case 'int':
                            val = _force_int(row[index])
                        case 'float':
                            val = _force_float(row[index])
                        case _:
                            val = row[index]
                    setattr(sub, attr_name, val)
                subs.append(sub)
        logging.info(f'Extracted {len(subs)} data rows.')
        return subs
    except Exception as e:
        logging.info(
            f"Exception encountered extracting values from {ws.title} for {key}", exc_info=e)
        return subs


def _build_apr(wb: Workbook, row: tuple, wb_map: dict) -> ESF_APR:
    # Create an instance of the ESF_APR class with the common
    # attributes for all APRs.
    apr = ESF_APR(omb_control_number=wb_map['omb_control_number'],
                  expiration_date=wb_map['expiration_date'],
                  cui_official=wb_map['cui_official'],
                  reporting_year=wb_map['reporting_year'])
    # Loop over the workbook map 'main" list to add all the attributes from the specific
    # APR subfund
    for attr in wb_map.get('main',_empty_gen()):
        index = attr.get('index', None)
        if index is None:
            logging.info(f'No index in APR map entry {attr}.')
            continue
        if index > len(row):
            logging.info(
                f'Index {index} exceeds row tuple length of {len(row)}')
            logging.info(f'row')
            continue
        attr_name = attr.get('name', None)
        if attr_name is None:
            logging.info(f'Missing attribute name in APR map enter {attr}')
            continue
        try:
            match attr.get('type', None):
                case 'bool':
                    val = _force_bool(row[index])
                case 'int':
                    val = _force_int(row[index])
                case 'float':
                    val = _force_float(row[index])
                case _:
                    val = row[index]
            setattr(apr, attr_name, val)
        except Exception as e:
            logging.error('Exception in _build_apr', exc_info=e)
            logging.error(row)
            logging.error(attr)
            logging.error(f'Index {index}')

    sublist = wb_map.get('subs',None)
    if sublist is not None:
        apr_key = None
        apr_key_name = wb_map.get('primary_grantee_key_name',None)
        if apr_key_name is None:
            logging.error('Missing primary grantee key name in workbook map')
        apr_key = getattr(apr,apr_key_name,None)
        if apr_key is None:
            logging.error(f'APR object missing key field name {apr_key_name} for extracting subaward records.')
            logging.error(f'{apr}')
        else:
            # Loop over all the subordinate pieces in the 'subs' list,
            # merging the contents of worksheets where needed and
            # creating a list of ESF_Sub instances for each.
            for sub in sublist:
                sub_name = sub.get('name',None)
                if sub_name is None:
                    logging.info('No name for subordinate list.')
                    continue
                merge_field = sub.get('merge_field',None)
                subvals = []
                sub_pieces = sub.get('children',_empty_gen())
                if len(sub_pieces) == 1:
                    ws = wb[sub_pieces[0].get('worksheet_name')]
                    field_map = sub_pieces[0].get('field_map')
                    key_offset = sub_pieces[0].get('key_offset')
                    subvals = _extract_sub_worksheet(ws=ws,
                        workbook_map=field_map, key=apr_key,
                        key_offset=key_offset)
                else:
                    for sub_piece in sub_pieces:
                        ws = wb[sub_piece.get('worksheet_name')]
                        field_map = sub_piece.get('field_map')
                        key_offset = sub_piece.get('key_offset')
                        partial_subvals = _extract_sub_worksheet(ws=ws,
                            workbook_map=field_map, key=apr_key,
                            key_offset=key_offset)
                        # Merge all the objects extracted from the child worksheet
                        # into the subvals list of dictionaries, using the key_field to
                        # determine whether there's an existing entry in the main list
                        # to update, or a new entry is needed.
                        for subval in partial_subvals:
                            merge_key = getattr(subval,merge_field, None)
                            if merge_key is None:
                                logging.error(f'Merge field {merge_field} missing in {subval}')
                                continue
                            else:
                                should_append = True
                                for val in subvals:
                                    merge_val = getattr(val,merge_field, None)
                                    if merge_key is not None and merge_key == merge_val:
                                        val.merge(subval)
                                        should_append = False
                                        break
                                if should_append:
                                    subvals.append(subval)
                setattr(apr, sub_name, subvals)

    return apr


class APRWorkbookList(Iterable):
    def __init__(self, wb: Workbook, config: dict):
        """Abstract iterating over any APR workbook."""
        self._key_iterator = None
        self._apr_iterator = None
        self._wb = wb
        self._config = config

    def __iter__(self):
        if self._config.get('primary_grantee_keys',None) is None:
            # Extract the worksheet named in the configuration.
            apr_ws = self._wb[self._config['primary_grantee_worksheet_name']]
            # Store an iterator for the entire APR worksheet.
            self._apr_iterator = apr_ws.iter_rows(min_row=2, min_col=1, max_col=len(self._config['main']), values_only=True)
        else:
            self._key_iterator = _key_gen(self._config['primary_grantee_keys'])
        return self

    def __next__(self):
        apri = None
        apr_row = None
        if self._key_iterator is not None:
            # Find the specific row matching the next key.
            key = next(self._key_iterator)
            apr_ws = self._wb[self._config['primary_grantee_worksheet_name']]
            key_column = self._config['primary_grantee_key_worksheet_column']
            row = _find_row(apr_ws, key=key, key_column=key_column)
            if row < 0:
                logging.error(f'Key {key} not found.')
                return None
            else:
                apr_iterator = apr_ws.iter_rows(min_row=row, max_row=row,
                    min_col=1, max_col=len(self._config['main']),
                    values_only=True)
                apr_row = next(apr_iterator)
        else:
            apr_row = next(self._apr_iterator)
        if apr_row is None:
            logging.error('No row retrieved from primary worksheet in APRWorkbookList.__next()__')
        apri = _build_apr(wb=self._wb, row=apr_row, wb_map=self._config)
        output_file_base_name = f"{self._config['subfund']}-{self._config['reporting_year']}"
        for fnc in self._config['filename_components']:
            component = getattr(apri,fnc,None)
            if component is not None:
                output_file_base_name = f'{output_file_base_name}-{component}'
        setattr(apri,'output_file_base_name',output_file_base_name)
        return apri
