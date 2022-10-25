# -*- coding: utf-8 -*-
"""
Created on Thu Aug 18 15:57:55 2022

@author: Keith.Tucker

Python command-line script for generating the Annual Performance Reports
(APR) for the Governor's Emergency Education Relief (GEER) Fund grants.

This program connects to an S3 bucket, finds the most reccent file
containing GEER APR data (or uses the file named on the command line)
and generates Portable Document Format (PDF) files for each record in the
file.

A second command line argument can be optionally specified to identify
a specific grantee for which a PDF should be generated, allowing an
update of just that grantee's APR.

Parameters for the session and service connection can be obtained from
the environment variables listed below.

     AWS_ACCESS_KEY_ID
     AWS_SECRET_ACCESS_KEY
     AWS_PROFILE
     AWS_DEFAULT_REGION

Alternatively, the parameters may be read in from configuration files
named [HOME]/.aws/credentials and [HOME}/.aws/config, where [HOME} is the
home directory for the account under which the program is run. See the URL
below for more information .

https://boto3.amazonaws.com/v1/documentation/api/latest/guide/credentials.html
"""

import argparse
import datetime
import logging
import os
import pathlib
import re
import sys

import boto3
import openpyxl

from geer_apr import GEER_Grantee, GEER_Subgrantee, GEER_APR

fileNamePattern = 'GEER Collection Data cleaned-'
defaultRegion = 'us-east-1'
defaultProfile = 'esf_dmp_published'

def find_latest_apr_file(bucket, fileName=None):
    latest = None
    try:
        if fileName:
            # Look up the modification time of the passed file.
            latest = bucket.Object(fileName).last_modified
        else:
            for bucketObj in bucket.objects.all():
                if latest is None or bucketObj.last_modified > latest:
                    if re.match(fileNamePattern,bucketObj.key):
                        fileName = bucketObj.key
                        latest = bucketObj.last_modified
    except Exception as e:
        logging.error(e)
    return fileName, latest

def find_grantee(ws, granteeID) -> int:
    for cell in ws['A']:
        if cell.value == granteeID:
            return cell.row
    # Fallback if grantee ID is not found
    return -1

def extract_subawards(worksheet, key, column) -> list:
    # Loop over the passed worksheet object, extracting all the values into
    # a list for the rows that match the passed key in the passed column.
    subs=[]
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[column] == key:
            sub = GEER_Subgrantee(
        		stateCode = row[0],
        		pk = row[1],
        		sk = row[2],
        		leaName = row[3],
        		iheName = row[4],
        		entityName = row[5],
        		dunsNumber = row[6],
        		awardedAmount = row[7],
        		awardedAmount_FLAG = row[8],
        		servedPopulation = row[9],
        		servedPopulation_FLAG = row[10],
        		fundsExpendedOnPublicSchools = row[11],
        		fundsExpendedOnPublicSchools_FLAG = row[12],
        		fundsExpendedOnNonPublicSchools = row[13],
        		fundsExpendedOnNonPublicSchools_FLAG = row[14],
        		fundsExpendedTotal = row[15],
        		fundsExpendedTotal_FLAG = row[16],
        		usedFundsForEducationalTechnology = row[17],
        		usedFundsForEducationalTechnology_FLAG = row[18],
        		usedFundsToAssistDisadvantaged = row[19],
        		usedFundsToAssistDisadvantaged_FLAG = row[20],
        		usedFundsForMentalHealth = row[21],
        		usedFundsForMentalHealth_FLAG = row[22],
        		usedFundsForSanitization = row[23],
        		usedFundsForSanitization_FLAG = row[24],
        		usedFundsForSummerAndAfterSchool = row[25],
        		usedFundsForSummerAndAfterSchool_FLAG = row[26],
        		usedFundsForOther = row[27],
        		usedFundsForOther_FLAG = row[28],
        		fundsUsedForOtherDescription = row[29],
        		fundsUsedForOtherDescription_FLAG = row[30],
        		usedFundsToProvideInternet = row[31],
        		usedFundsToProvideInternet_FLAG = row[32],
        		usedFundsForMobileHotspots = row[33],
        		usedFundsForMobileHotspots_FLAG = row[34],
        		usedFundsForInternetDevices = row[35],
        		usedFundsForInternetDevices_FLAG = row[36],
        		usedFundsForHomeInternet = row[37],
        		usedFundsForHomeInternet_FLAG = row[38],
        		usedFundsForDistrictInternet = row[39],
        		usedFundsForDistrictInternet_FLAG = row[40],
        		usedFundsForOtherInternet = row[41],
        		usedFundsForOtherInternet_FLAG = row[42],
        		usedFundsForOtherInternetDescription = row[43],
        		usedFundsForOtherInternetDescription_FLAG = row[44],
        		usedFundsForDedicatedLearningDevices = row[45],
        		usedFundsForDedicatedLearningDevices_FLAG = row[46],
        		elementaryStudentsWithDedicatedDevice = row[47],
        		elementaryStudentsWithDedicatedDevice_FLAG = row[48],
        		elementaryStudentsEnrolled = row[49],
        		elementaryStudentsEnrolled_FLAG = row[50],
        		elementaryStudentProportionWithDevices = row[51],
        		elementaryStudentProportionWithDevices_FLAG = row[52],
        		elementaryStudentProportionWithDevices_calc = row[53],
        		elementaryStudentProportionWithDevices_calc_FLAG = row[54],
        		secondaryStudentsWithDedicatedDevice = row[55],
        		secondaryStudentsWithDedicatedDevice_FLAG = row[56],
        		secondaryStudentsEnrolled = row[57],
        		secondaryStudentsEnrolled_FLAG = row[58],
        		secondaryStudentProportionWithDevices = row[59],
        		secondaryStudentProportionWithDevices_FLAG = row[60],
        		secondaryStudentProportionWithDevices_calc = row[61],
        		secondaryStudentProportionWithDevices_calc_FLAG = row[62],
        		fundsExpendedByIhe = row[63],
        		fundsExpendedByIhe_FLAG = row[64],
        		fundsUsedToProvideFinancialAid = row[65],
        		fundsUsedToProvideFinancialAid_FLAG = row[66],
        		numberOfStudentsReceivedFinancialAid = row[67],
        		numberOfStudentsReceivedFinancialAid_FLAG = row[68],
        		fundsExpendedByEntity = row[69],
        		fundsExpendedByEntity_FLAG = row[70],
        		isPreKServed = row[71],
        		isPreKServed_FLAG = row[72],
        		isK12Served = row[73],
        		isK12Served_FLAG = row[74],
        		isPostSecServed = row[75],
        		isPostSecServed_FLAG = row[76],
        		isDistanceLearningSupported = row[77],
        		isDistanceLearningSupported_FLAG = row[78],
        		isDirectFinancialSupportProvided = row[79],
        		isDirectFinancialSupportProvided_FLAG = row[80],
        		ftePositionsAsOf09302018 = row[81],
        		ftePositionsAsOf09302018_FLAG = row[82],
        		ftePositionsAsOf09302019 = row[83],
        		ftePositionsAsOf09302019_FLAG = row[84],
        		ftePositionsAsOf03132020 = row[85],
        		ftePositionsAsOf03132020_FLAG = row[86],
        		ftePositionsAsOf09302020 = row[87],
        		ftePositionsAsOf09302020_FLAG = row[88],
            )
            subs.append(sub)
    return subs
            
    
def generate_aprs(bucket, fileName, latest, granteeID=None):
    try:
        # If the file does not already exist or is older than the copy found in
        # cloud storage, download it for local use.
        download = False
        try:
            fstat = os.stat(fileName)
            if datetime.datetime.fromtimestamp(fstat.st_mtime,
                                               tz=datetime.timezone.utc) < latest:
                download = True
        except os.error:
            # The file must not exist, so download it.
            download = True
        if download:
            # Download the object as a local file.
            with open(fileName, 'wb') as f:
                bucket.download_fileobj(fileName, f)

        # Read the file as an Excel file, iterating over the rows to generate
        # PDFs from the column content.
        wb = openpyxl.load_workbook(filename=fileName,read_only=True, data_only=True)
        ws = wb["Prime Awards"]
        if granteeID:
            # Find the row containing the grantee ID
            row = find_grantee(ws, granteeID)
            # Generate the PDF for just that grantee ID
            if row >= 0:
                apr = GEER_APR()
                subs=extract_subawards(wb["Sub Grantees"], granteeID, 0)
                grantee=GEER_Grantee(
                       stateCode = ws.cell(row,1).value,
                       pk = ws.cell(row,2).value,
                       sk = ws.cell(row,3).value,
                       prNumber = ws.cell(row,4).value,
                       lastModifiedBy = ws.cell(row,5).value,
                       lastModifiedDate = ws.cell(row,6).value,
                       submittedBy = ws.cell(row,7).value,
                       submittedDate = ws.cell(row,8).value,
                       reportingYear = ws.cell(row,9).value,
                       granteeName = ws.cell(row,10).value,
                       repName = ws.cell(row,11).value,
                       repPosition = ws.cell(row,12).value,
                       repOffice = ws.cell(row,13).value,
                       repMailingAddress = ws.cell(row,14).value,
                       repPhone = ws.cell(row,15).value,
                       repEmail = ws.cell(row,16).value,
                       grantAmountAllocated = ws.cell(row,17).value,
                       grantAmountExpended = ws.cell(row,18).value,
                       awardedAmount_sub = ws.cell(row,19).value,
                       awardedAmount_sub_FLAG = ws.cell(row,20).value,
                       areLeasAwardedGeerFunds = ws.cell(row,21).value,
                       areLeasAwardedGeerFunds_FLAG = ws.cell(row,22).value,
                       areIhesAwardedGeerFunds = ws.cell(row,23).value,
                       areIhesAwardedGeerFunds_FLAG = ws.cell(row,24).value,
                       areEntitiesAwardedGeerFunds = ws.cell(row,25).value,
                       areEntitiesAwardedGeerFunds_FLAG = ws.cell(row,26).value,
                       wereAnyConditionsPlacedByStateForLeaFunds = ws.cell(row,27).value,
                       isStateLeaGeerAwardConditionChanges = ws.cell(row,28).value,
                       stateLeaGeerAwardConditionChanges = ws.cell(row,29).value,
                       didStatePlaceDistanceLearningConditionsOnLeas = ws.cell(row,30).value,
                       isSupportTechInfrastructureForLeaDistanceLearning = ws.cell(row,31).value,
                       isInternetAccessNeededForLeas = ws.cell(row,32).value,
                       areDevicesNeededForLeas = ws.cell(row,33).value,
                       isTrainingStaffNeededForLeas = ws.cell(row,34).value,
                       provideDigitalLearningContentForLeas = ws.cell(row,35).value,
                       areOtherConditionsForLeas = ws.cell(row,36).value,
                       otherConditionsForLeas = ws.cell(row,37).value,
                       wereAnyConditionsPlacedByStateForIheFunds = ws.cell(row,38).value,
                       isStateIheGeerAwardConditionChanges = ws.cell(row,39).value,
                       stateIheGeerAwardConditionChanges = ws.cell(row,40).value,
                       didStatePlaceDistanceLearningConditionsOnIhes = ws.cell(row,41).value,
                       isSupportTechInfrastructureForIheDistanceLearning = ws.cell(row,42).value,
                       isInternetAccessNeededForIhes = ws.cell(row,43).value,
                       areDevicesNeededForIhes = ws.cell(row,44).value,
                       isTrainingStaffNeededForIhes = ws.cell(row,45).value,
                       provideDigitalLearningContentForIhes = ws.cell(row,46).value,
                       areOtherConditionsForIhes = ws.cell(row,47).value,
                       otherConditionsForIhes = ws.cell(row,48).value,
                       didStateDirectAnyIhesToUseGeerFundsForEmergency = ws.cell(row,49).value,
                       numberOfPublicSchoolsReceivedGeerFunds = ws.cell(row,50).value,
                       numberOfNonPublicSchoolsReceivedGeerFunds = ws.cell(row,51).value
                )
                apr.generate(grantee,subs)
            else:
                logging.error(f'Grantee ID {granteeID} not found.')
        else:
            for row in ws.iter_rows(min_row=2, values_only=True):
                apr = GEER_APR()
                grantee=GEER_Grantee(
                       stateCode = row[0],
                       pk = row[1],
                       sk = row[2],
                       prNumber = row[3],
                       lastModifiedBy = row[4],
                       lastModifiedDate = row[5],
                       submittedBy = row[6],
                       submittedDate = row[7],
                       reportingYear = row[8],
                       granteeName = row[9],
                       repName = row[10],
                       repPosition = row[11],
                       repOffice = row[12],
                       repMailingAddress = row[13],
                       repPhone = row[14],
                       repEmail = row[15],
                       grantAmountAllocated = row[16],
                       grantAmountExpended = row[17],
                       awardedAmount_sub = row[18],
                       awardedAmount_sub_FLAG = row[19],
                       areLeasAwardedGeerFunds = row[20],
                       areLeasAwardedGeerFunds_FLAG = row[21],
                       areIhesAwardedGeerFunds = row[22],
                       areIhesAwardedGeerFunds_FLAG = row[23],
                       areEntitiesAwardedGeerFunds = row[24],
                       areEntitiesAwardedGeerFunds_FLAG = row[25],
                       wereAnyConditionsPlacedByStateForLeaFunds = row[26],
                       isStateLeaGeerAwardConditionChanges = row[27],
                       stateLeaGeerAwardConditionChanges = row[28],
                       didStatePlaceDistanceLearningConditionsOnLeas = row[29],
                       isSupportTechInfrastructureForLeaDistanceLearning = row[30],
                       isInternetAccessNeededForLeas = row[31],
                       areDevicesNeededForLeas = row[32],
                       isTrainingStaffNeededForLeas = row[33],
                       provideDigitalLearningContentForLeas = row[34],
                       areOtherConditionsForLeas = row[35],
                       otherConditionsForLeas = row[36],
                       wereAnyConditionsPlacedByStateForIheFunds = row[37],
                       isStateIheGeerAwardConditionChanges = row[38],
                       stateIheGeerAwardConditionChanges = row[39],
                       didStatePlaceDistanceLearningConditionsOnIhes = row[40],
                       isSupportTechInfrastructureForIheDistanceLearning = row[41],
                       isInternetAccessNeededForIhes = row[42],
                       areDevicesNeededForIhes = row[43],
                       isTrainingStaffNeededForIhes = row[44],
                       provideDigitalLearningContentForIhes = row[45],
                       areOtherConditionsForIhes = row[46],
                       otherConditionsForIhes = row[47],
                       didStateDirectAnyIhesToUseGeerFundsForEmergency = row[48],
                       numberOfPublicSchoolsReceivedGeerFunds = row[49],
                       numberOfNonPublicSchoolsReceivedGeerFunds = row[50]
                )
                subs=extract_subawards(wb["Sub Grantees"],grantee.stateCode,0)
                apr.generate(grantee,subs)
            
    except Exception as e:
        logging.error(e)
    
if __name__ == '__main__':

    logging.basicConfig(level=os.environ.get("LOGLEVEL",logging.ERROR))
    
    # Retrieve the parameters from environment variables, if set.
    access_key_id = os.getenv('AWS_ACCESS_KEY_ID')
    secret_access_key = os.getenv('AWS_SECRET_ACCESS_KEY')
    profile = os.getenv('AWS_PROFILE', defaultProfile)
    region = os.getenv('AWS_DEFAULT_REGION', defaultRegion)

    try:
        session = boto3.Session(aws_access_key_id=access_key_id,
                                aws_secret_access_key=secret_access_key,
                                profile_name=profile,
                                region_name=region)
        s3 = session.resource('s3')
        bucket = s3.Bucket('prod-esf-dmp-published')

        ap = argparse.ArgumentParser(formatter_class=argparse.RawDescriptionHelpFormatter,
        description='''Generate Annual Performance Reports for the GEER grants.

  Unless overridden on the command line, APRs for all grantees are generated
  in the current directory, using the latest version of the APR information collection instruments,
  and using data in the latest file in cloud storage matching a file naming pattern.
''',
        epilog='''The program uses defaults for cloud storage access, which can be overridden with the following environment variables:
  AWS_ACCESS_KEY_ID: The identifier for authenticating access to cloud storage
  AWS_SECRET_ACCESS_KEY: The associated access key for authenticating access to cloud storage
  AWS_PROFILE: The name of the cloud storage location
  AWS_DEFAULT_REGION: The name of the associated Cloud Service Provider region in which the cloud storage is located
''')
        ap.add_argument('-f','--filename', help='Use the data in the specified file.') 
        ap.add_argument('-g','--grantee', help='Generate for just the specific grantee.')
        ap.add_argument('-v','--version', help='Generate the specific version of the APR.')
        ap.add_argument('-o','--output',help='Generate in the specified directory.')
        args = ap.parse_args()
        if args.output:
            outdir = pathlib.Path(args.output).resolve(strict=False)
            if outdir.exists():
                if outdir.is_dir:
                    os.chdir(outdir)
                else:
                    logging.error(f'{args.output} exists, but is not a directory. --output parameter must be a directory name.')
                    exit()
            else:
                outdir.mkdir(parents = True, exist_ok = True)
                os.chdir(outdir)
        fileName, mtime = find_latest_apr_file(bucket, args.filename)

        generate_aprs(bucket,fileName,mtime,args.grantee)

    except Exception as e:
        logging.error(e)
